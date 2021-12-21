# -*- coding:utf-8 -*-
# Author:toby
# Date : 2021/12/17 14:01
"""
一、一个完整的闭包必须满足哪几个条件
1、函数中嵌套一个函数
2、外层函数返回内层嵌套函数的函数名
3、内层嵌套函数有引用外层的一个非全局变量


二、定义一个计算函数运行时间的装饰器(计算时间使用time模块实现)
3、编写装饰器，为多个函数加上认证的功能(用户的账号密码来自文件)，要求登录成功一次，后续的函数都无须再输入账号和密码
"""
from openpyxl import load_workbook

import time

print("--------------------计算函数运行时间--------------------")


# 装饰器 <== 高阶函数 + 嵌套函数
def wrapper(func):  # 计算函数运行时间
    def count_time(*args, **kwargs):
        start_time = time.time()
        print("开始时间", start_time)
        func(*args, **kwargs)
        time.sleep(1)
        end_tine = time.time()
        print("结束时间", end_tine)
        print("函数运行的时间为:{:.5f}s".format(end_tine - start_time))

    return count_time  # 这里的'()'会导致报错'NoneType' object is not callable
    # 只要去掉count_time后面的括号即可解决问题


@wrapper  # 引用装饰器，等同于 add_num = wrapper(add_num)
def add_num(a, b):
    return a + b


add_num(22, 33)

print("--------------------计算函数运行时间--------------------")
print("--------------------读取excel--------------------")


# 读取每一条测试用用例分别保存到字典中，然后再将所有用例保存到列表中，如[{用例1},{用例2},{用例3}]
class DoExcel:
    @staticmethod
    def read_case(file_path, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        row_max = sheet.max_row  # 最大行
        column_max = sheet.max_column  # 最大列
        test_case = []
        for row in range(2, row_max + 1):
            user_data = {'username': sheet.cell(row, 1).value,
                         'password': sheet.cell(row, 2).value,
                         'token': sheet.cell(row, 3).value}
            test_case.append(user_data)
        return test_case

    @staticmethod  # 静态方法
    def write_back(file_name, sheet_name, i, j, token):
        """
        :param file_name: 文件名
        :param sheet_name: 表名
        :param i: 单元格位置,行
        :param j: 单元格位置,列
        :param token: token值
        :return:
        """
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, j).value = token
        wb.save(file_name)  # 保存结果


if __name__ == '__main__':
    test_data = DoExcel.read_case('testdata.xlsx', 'user')
    DoExcel.write_back('testdata.xlsx', 'user', 2, 3, 'True')
    print(test_data)

    print("--------------------读取excel--------------------")
