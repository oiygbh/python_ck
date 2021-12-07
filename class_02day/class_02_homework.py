# -*- coding:utf-8 -*-
# Author:toby
# Date : 2021/12/7 9:19
from openpyxl import load_workbook

#  1、使用字典推倒是将下面字符串格式的数据，改成字典类型的数据

cook_str = 'BIDUPSID=D0727533D7147B7;PSTM=1530348042;BAIDUID=81005C9BC2EB28;' \
           'sugstore=0;__ cfdui d=d0a13458f8ac2a;' \
           'BD _UPN=12314353;ispeed _1sm=2 ;BDORZ=B490B5EBF6F3CD402'

dict_str = {i.split("=")[0]: i.split("=")[1] for i in cook_str.split(";")}
print(dict_str)


#  2、当前有文件case.exce1,设计程序将exce1中的用例读取到一个生成器?

def read_excel(file_path, sheet_name):
    """

    读取excel数据的方法
    :param file_path: 文件路径
    :param sheet_name: 表名
    :return: 读取数据
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    row_max = sheet.max_row  # 最大行
    column_max = sheet.max_column  # 最大列
    for row in range(2, row_max + 1):
        row_dict = {}  # 每行数据
        for col in range(1, column_max + 1):
            # key等于表头具体值
            row_dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col)
        yield row_dict
#  3、请描述什么是可迭代对象?什么是迭代器? 迭代器和生成器的区别?
# 可迭代对象：
# 1、可以使用for循环遍历的对象，如list、tuple、dict、set、str
# 2、生成器（Generator）
# 可以使用isinstance()来判断一个对象是否是Iterable对象：
# from collections import Iterable
# print(isinstance([], Iterable))
# 迭代器：
# 可以被next()函数调用并不断返回下一个值的对象称为迭代器：Iterator。
# 迭代器和生成器区别：
# 生成器能做到迭代器能做的所有事,而且因为自动创建了 iter()和 next()方法,生成器显得特别简洁,而且
# 生成器也是高效的，使用生成器表达式取代列表解析可以同时节省内存。除了创建和保存程序状态的自动方法,当
# 发生器终结时,还会自动抛出 StopIteration异常
